import openpyxl

file_path = "/Users/xh/School/Final Year/cyclerap/PathSafetyAssessmentTool/CycleRAP - model - generation v2.11 - 24.09.27 - suppliers.xlsm"
search_term = "Upgrade to on-road bicycle lane"

print(f"Loading '{file_path}'...")
wb = openpyxl.load_workbook(file_path, read_only=True, keep_vba=True)

print(f"Searching for '{search_term}' in all sheets...")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Scanning sheet '{sheet_name}'...")
    found = False
    try:
        # Search first 100 rows and 20 columns
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and search_term in cell.value:
                    print(f"FOUND in '{sheet_name}' at {cell.coordinate}: {cell.value}")
                    found = True
                    break
            if found: break
    except Exception as e:
        print(f"Error scanning sheet: {e}")
