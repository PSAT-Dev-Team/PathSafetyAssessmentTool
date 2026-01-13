import openpyxl

file_path = "/Users/xh/School/Final Year/cyclerap/PathSafetyAssessmentTool/CycleRAP - model - generation v2.11 - 24.09.27 - suppliers.xlsm"
sheet_name = "CycleRAP Model"

print(f"Loading '{file_path}'...")
# data_only=False to read formulas
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False, keep_vba=True)
ws = wb[sheet_name]

print(f"Inspecting formulas in '{sheet_name}' (Rows 13-15)...")

# Based on previous output:
# VB-severity is Col 123 (DS)
# VB-CF is Col 124 (DT)
cols_to_inspect = [123, 124] 

for row_idx in range(13, 16):
    print(f"\nRow {row_idx}:")
    for col_idx in cols_to_inspect:
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        # If it starts with =, it's a formula
        if isinstance(val, str) and val.startswith("="):
             print(f"  Col {col_idx} ({cell.coordinate}) HEADER?: {ws.cell(row=2, column=col_idx).value} -> FORMULA: {val}")
        elif val is not None:
             print(f"  Col {col_idx} ({cell.coordinate}) HEADER?: {ws.cell(row=2, column=col_idx).value} -> VALUE: {val}")
